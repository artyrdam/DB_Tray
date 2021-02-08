import sys
import openpyxl
from tkinter import filedialog as fd


def get_excel_data():
    path = fd.askopenfilename()
    if path == "":
        exit()

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    data_list = []
    # prints all objects in column 1
    for i in range(1, sheet_obj.max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        # print(cell_obj.value)
        data_list.append(cell_obj.value)

    # remove dupes
    data_list = list(dict.fromkeys(data_list))

    x = path.split(".")[0]
    new_csv_file = x + ".csv"
    # print("path: " + new_csv_file)

    with open(new_csv_file, 'w') as myfile:
        myfile.write("\n".join(data_list))


if __name__ == "__main__":
    get_excel_data()
